import openpyxl
import os
from openpyxl import Workbook, load_workbook
import pandas as pd
import matplotlib.pyplot as plt

def prepare_data(file_path):
    wiersze = 101  # w jednej zakladce 100 wiersz dla plci
    kolumny = 17    #liczba punktow
    tab_m = [[None] * kolumny for _ in range(wiersze)]
    tab_k = [[None] * kolumny for _ in range(wiersze)]

    wb = openpyxl.load_workbook(file_path)

    #liczba zakladek
    liczba_zakladek = 17 #(od 2006)

    # Iteracja przez zakładki
    for i in range(2006, 2006 + liczba_zakladek):
        # Składanie nazwy zakładki
        nazwa_zakladki = f'{i}'

        # Wybór zakładki
        sheet = wb[nazwa_zakladki]

        # Przetwarzanie danych dla mężczyzn
        for j in range(5, 106):
            wiersz = j
            kolumna = 'D'
            wartosc = sheet[f'{kolumna}{wiersz}'].value
            if wartosc is not None:
                tab_m[j - 5][i - 2006] = 1 - wartosc

        # Przetwarzanie danych dla kobiet
        for j in range(106, 207):
            wiersz = j
            kolumna = 'D'
            wartosc = sheet[f'{kolumna}{wiersz}'].value
            if wartosc is not None:
                tab_k[j - 106][i - 2006] = 1 - wartosc

    # Debugowanie - drukowanie wartości w tab_m i tab_k
    # print("Dane mężczyzn (tab_m):")
    # for row in tab_m:
    #     print(row)
    #
    # print("Dane kobiet (tab_k):")
    # for row in tab_k:
    #     print(row)

    return tab_m, tab_k


def dane_wykresy(macierz):
    wiersze = len(macierz)
    kolumny = len(macierz[0])
    dane_wykres = []

    # Od prawego górnego rogu do lewego dolnego rogu
    for k in range(kolumny - 1, -1, -1):
        dane = []
        i, j = 0, k
        while j < kolumny and i < wiersze:
            if macierz[i][j] is not None:  # Ignorowanie wartości None
                dane.append(macierz[i][j])
            i += 1
            j += 1
        print(f"Przekątna od (0, {k}) do końca: {dane}")
        if dane:  # Dodawanie tylko niepustych list
            dane_wykres.append(dane)

    # Od drugiego wiersza do ostatniego
    for k in range(1, wiersze):
        dane = []
        i, j = k, 0
        while i < wiersze and j < kolumny:
            if macierz[i][j] is not None:  # Ignorowanie wartości None
                dane.append(macierz[i][j])
            i += 1
            j += 1
        print(f"Przekątna od ({k}, 0) do końca: {dane}")
        if dane:  # Dodawanie tylko niepustych list
            dane_wykres.append(dane)

    return dane_wykres




def save_data_to_excel(file_path_men, file_path_women, file_path_a, tab_m, tab_k):
    # Zapisanie danych dla mężczyzn
    wb_m = Workbook()
    ws_m = wb_m.active
    ws_m.title = "Mężczyźni"

    # Dodaj nagłówki kolumn (punkty)
    years = list(range(0, 17))
    ws_m.append(["punkt "] + years)

    # Dodaj wiersze z wiekiem (0-100)
    for year in range(2022,1905, -1):
        ws_m.append([year])
    

    # Dodaj dane do arkusza "Mężczyźni"
    dane_do_wykresu_M = dane_wykresy(tab_m)
    for row_index, row_data in enumerate(dane_do_wykresu_M, start=2):
        probability = 1
        for col_index, value in enumerate(row_data, start=2):
            if value is not None:
                probability *= value
            ws_m.cell(row=row_index, column=col_index, value=probability * 100)

    wb_m.save(file_path_men)
    # print(f"Dane zostały zapisane do pliku {file_path_men}")

    # Zapisanie danych dla kobiet
    wb_k = Workbook()
    ws_k = wb_k.active
    ws_k.title = "Kobiety"

    # Dodaj nagłówki kolumn (lata)
    ws_k.append(["punkt"] + years)

    # Dodaj wiersze z wiekiem (0-100)
    for age in range(2022, 1905, -1):
        ws_k.append([age])
    

    # Dodaj dane do arkusza "Kobiety"
    dane_do_wykresu_K = dane_wykresy(tab_k)
    for row_index, row_data in enumerate(dane_do_wykresu_K, start=2):
        probability = 1
        for col_index, value in enumerate(row_data, start=2):
            if value is not None:
                probability *= value
            ws_k.cell(row=row_index, column=col_index, value=probability * 100)

    wb_k.save(file_path_women)

    wb_a = Workbook()
    ws_a = wb_a.active
    ws_a.title = "Ogólne"

    # Dodaj nagłówki kolumn (punkty)
    years = list(range(0, 17))
    ws_a.append(["punkt "] + years)

    # Dodaj wiersze z wiekiem (0-100)
    for year in range(2022, 1905, -1):
        ws_a.append([year])

    wb_m = load_workbook(file_path_men)
    ws_m = wb_m.active

    wb_k = load_workbook(file_path_women)
    ws_k = wb_k.active

    # Dodaj dane do arkusza "Średnia"
    for row in range(2, ws_m.max_row + 1):
        for col in range(2, ws_m.max_column + 1):
            value_m = ws_m.cell(row=row, column=col).value
            value_k = ws_k.cell(row=row, column=col).value

            # Oblicz średnią, uwzględniając brakujące dane
            if value_m is not None and value_k is not None:
                avg_value = (value_m + value_k) / 2
            elif value_m is not None:
                avg_value = value_m
            elif value_k is not None:
                avg_value = value_k
            else:
                avg_value = None

            ws_a.cell(row=row, column=col, value=avg_value)
            print(f"Średnia - Wiersz: {row}, Kolumna: {col}, Wartość: {avg_value}")
    wb_a.save(file_path_a)
    print(f"Dane zostały zapisane do pliku {file_path_a}")

# # Główna część programu
# file_path = 'tablice_trwania_zycia_w_latach_1990-2022.xlsx'
# file_path_men = 'preprocessed_male.xlsx'
# file_path_women = 'preprocessed_female.xlsx'
#
# # Przygotowanie danych
# tab_m, tab_k = prepare_data(file_path)
#
# # Zapisanie danych do osobnych plików Excel
# save_data_to_excel(file_path_men, file_path_women, tab_m, tab_k)

def lineChartOne(sex, year):
    if sex == 0:
        path = 'data/dane_mezczyzni.xlsx'
    if sex == 1:
        path = 'data/dane_kobiety.xlsx'
    if sex == 2:
        path = 'data/dane_ogolne.xlsx'

    data = pd.read_excel(path)
    row = data[data.iloc[:, 0] == year]

    if row.empty:
        raise ValueError(f"Wrong year {year}.")

    # Extract the row data from the second column to the end
    row_data = row.iloc[0, 1:].dropna()

    row_data = pd.concat([pd.Series([100]), row_data]).reset_index(drop=True)

    # Plot the line chart
    fig, ax = plt.subplots(figsize=(10, 6))
    ax.plot(row_data.index, row_data.values, marker='o')
    ax.set_xlabel("Years")
    ax.set_ylabel("Percentage")
    ax.grid(True)

    # Return the figure object
    return fig

# test = lineChartOne(1, 1960)
# test.show()
def lineChartRange(sex, start, end):
    if sex == 0:
        path = 'data/dane_mezczyzni.xlsx'
    if sex == 1:
        path = 'data/dane_kobiety.xlsx'
    if sex == 2:
        path = 'data/dane_ogolne.xlsx'

    data = pd.read_excel(path)

    rows = data[(data.iloc[:, 0] >= start) & (data.iloc[:, 0] <= end)]

    if rows.empty:
        raise ValueError(f"Wrong year range: {start}-{end}.")

    avg_data = rows.iloc[:, 1:].apply(pd.to_numeric, errors='coerce').mean(skipna=True)

    avg_data = pd.concat([pd.Series([100]), avg_data]).reset_index(drop=True)
    #print(avg_data)

    fig, ax = plt.subplots(figsize=(10, 6))
    ax.plot(avg_data.index, avg_data.values, marker='o')
    ax.set_xlabel("Years")
    ax.set_ylabel("Percentage")
    ax.grid(True)


    return fig

# test2 = lineChartRange(1, 1960, 1970)
# test2.show()
